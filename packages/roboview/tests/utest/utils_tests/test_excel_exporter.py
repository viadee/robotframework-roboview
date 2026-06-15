"""Tests for Excel report exporter."""

import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from roboview.schemas.domain.reports import (
    SummaryReport,
    KPISummary,
    ReportMetadata,
    ReportTypeEnum,
    Recommendation,
)
from roboview.utils.exporters.excel_exporter import ExcelExporter


def test_excel_export_summary_report() -> None:
    """Test exporting summary report to Excel."""
    report = SummaryReport(
        metadata=ReportMetadata(
            project_name="Test Project",
            project_root="/test/project",
            author="Test Author",
        ),
        title="Summary Report",
        report_type=ReportTypeEnum.SUMMARY,
        summary=KPISummary(
            total_keywords=42,
            unused_keywords=3,
            reusage_rate=65.5,
            documentation_coverage=80.0,
            robocop_issues=12,
            total_files=15,
        ),
        risk_level="LOW",
        health_status="Good project health.",
        robocop_issues_by_category={"Documentation": 5, "Naming": 7},
    )

    with tempfile.TemporaryDirectory() as temp_dir:
        output_file = Path(temp_dir) / "report.xlsx"
        ExcelExporter.export(report, output_file)

        # Verify file exists
        assert output_file.exists()

        # Verify it's a valid Excel file
        wb = load_workbook(output_file)
        assert len(wb.sheetnames) > 0


def test_excel_export_creates_all_sheets() -> None:
    """Test that Excel export creates all expected sheets."""
    report = SummaryReport(
        metadata=ReportMetadata(
            project_name="Test",
            project_root="/test",
        ),
        title="Multi-Sheet Report",
        report_type=ReportTypeEnum.SUMMARY,
        summary=KPISummary(
            total_keywords=30,
            unused_keywords=2,
            reusage_rate=70.0,
            documentation_coverage=85.0,
            robocop_issues=8,
            total_files=12,
        ),
        risk_level="LOW",
        health_status="Good project health.",
        robocop_issues_by_category={"Documentation": 3, "Naming": 5},
        recommendations=[
            Recommendation(
                priority="HIGH",
                category="Documentation",
                message="Add documentation",
            ),
        ],
    )

    with tempfile.TemporaryDirectory() as temp_dir:
        output_file = Path(temp_dir) / "multi_sheet.xlsx"
        ExcelExporter.export(report, output_file)

        wb = load_workbook(output_file)
        sheet_names = wb.sheetnames

        # Check for expected sheets
        assert "Metadata" in sheet_names
        assert "KPI Summary" in sheet_names
        assert "Robocop Issues" in sheet_names
        assert "Recommendations" in sheet_names


def test_excel_export_metadata_sheet_content() -> None:
    """Test that metadata sheet contains correct data."""
    author = "John Doe"
    project_name = "Test Project"
    
    report = SummaryReport(
        metadata=ReportMetadata(
            project_name=project_name,
            project_root="/test/project",
            author=author,
        ),
        title="Metadata Test",
        report_type=ReportTypeEnum.SUMMARY,
        summary=KPISummary(
            total_keywords=20,
            unused_keywords=1,
            reusage_rate=75.0,
            documentation_coverage=90.0,
            robocop_issues=5,
            total_files=10,
        ),
        risk_level="LOW",
        health_status="Good project health.",
        robocop_issues_by_category={},
    )

    with tempfile.TemporaryDirectory() as temp_dir:
        output_file = Path(temp_dir) / "metadata.xlsx"
        ExcelExporter.export(report, output_file)

        wb = load_workbook(output_file)
        ws = wb["Metadata"]

        # Check that metadata is present
        content = str(ws.values)
        assert project_name in content
        assert author in content


def test_excel_export_kpi_values() -> None:
    """Test that KPI sheet contains correct values."""
    report = SummaryReport(
        metadata=ReportMetadata(
            project_name="KPI Test",
            project_root="/test",
        ),
        title="KPI Values",
        report_type=ReportTypeEnum.SUMMARY,
        summary=KPISummary(
            total_keywords=100,
            unused_keywords=5,
            reusage_rate=80.0,
            documentation_coverage=95.0,
            robocop_issues=3,
            total_files=20,
        ),
        risk_level="OPTIMAL",
        health_status="Excellent project health.",
        robocop_issues_by_category={},
    )

    with tempfile.TemporaryDirectory() as temp_dir:
        output_file = Path(temp_dir) / "kpi.xlsx"
        ExcelExporter.export(report, output_file)

        wb = load_workbook(output_file)
        ws = wb["KPI Summary"]

        # Check for KPI values in sheet
        content = str(ws.values)
        assert "100" in content  # total_keywords
        assert "5" in content    # unused_keywords
        assert "80" in content   # reusage_rate


def test_excel_export_creates_directories() -> None:
    """Test that Excel export creates necessary directories."""
    report = SummaryReport(
        metadata=ReportMetadata(
            project_name="Dir Test",
            project_root="/test",
        ),
        title="Directory Test",
        report_type=ReportTypeEnum.SUMMARY,
        summary=KPISummary(
            total_keywords=15,
            unused_keywords=1,
            reusage_rate=80.0,
            documentation_coverage=90.0,
            robocop_issues=2,
            total_files=8,
        ),
        risk_level="OPTIMAL",
        health_status="Excellent project health.",
        robocop_issues_by_category={},
    )

    with tempfile.TemporaryDirectory() as temp_dir:
        output_file = Path(temp_dir) / "reports" / "2024" / "report.xlsx"
        ExcelExporter.export(report, output_file)

        assert output_file.exists()
        assert output_file.parent.exists()
